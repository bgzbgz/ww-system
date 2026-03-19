import pytest
import os
from services.excel_parser import parse_excel, ExcelParseError

FIXTURES = os.path.join(os.path.dirname(__file__), "fixtures")

def test_parse_valid_file_returns_correct_rows():
    rows = parse_excel(os.path.join(FIXTURES, "valid.xlsx"))
    assert len(rows) == 3
    assert rows[0]["full_name"] == "María García"
    assert rows[0]["first_name"] == "María"
    assert rows[0]["email"] == "m.garcia@banco.com"
    assert rows[0]["company"] == "Banco Agrícola"
    assert rows[0]["position"] == "Director"

def test_parse_derives_first_name_from_full_name():
    rows = parse_excel(os.path.join(FIXTURES, "valid.xlsx"))
    assert rows[1]["first_name"] == "Carlos"

def test_parse_missing_email_sets_empty_string_and_warning():
    rows = parse_excel(os.path.join(FIXTURES, "valid.xlsx"))
    assert rows[2]["email"] == ""
    assert rows[2]["warning"] == "missing_email"

def test_parse_invalid_email_sets_warning(tmp_path):
    import openpyxl
    path = tmp_path / "invalid_email.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Full Name", "Email", "Company"])
    ws.append(["Test Person", "not-an-email", "TestCo"])
    wb.save(path)
    rows = parse_excel(str(path))
    assert rows[0]["warning"] == "invalid_email"

def test_parse_skips_blank_rows(tmp_path):
    import openpyxl
    path = tmp_path / "blank_rows.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Full Name", "Email", "Company"])
    ws.append(["María García", "m@banco.com", "Banco"])
    ws.append([None, None, None])  # blank row
    ws.append(["Carlos López", "c@siman.com", "Siman"])
    wb.save(path)
    rows = parse_excel(str(path))
    assert len(rows) == 2

def test_parse_column_matching_is_case_insensitive(tmp_path):
    import openpyxl
    path = tmp_path / "case.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["FULL NAME", "EMAIL", "COMPANY"])
    ws.append(["Test Person", "test@example.com", "TestCo"])
    wb.save(path)
    rows = parse_excel(str(path))
    assert rows[0]["full_name"] == "Test Person"

def test_parse_raises_on_missing_required_column():
    with pytest.raises(ExcelParseError) as exc:
        parse_excel(os.path.join(FIXTURES, "missing_col.xlsx"))
    assert "full name" in str(exc.value).lower()
    assert "found columns" in str(exc.value).lower()

def test_parse_flags_duplicate_names(tmp_path):
    import openpyxl
    path = tmp_path / "dupes.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Full Name", "Email", "Company"])
    ws.append(["María García", "a@a.com", "CoA"])
    ws.append(["María García", "b@b.com", "CoB"])
    wb.save(path)
    rows = parse_excel(str(path))
    assert rows[0]["warning"] == "duplicate_name"
    assert rows[1]["warning"] == "duplicate_name"
