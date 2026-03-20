import re
from collections import Counter
import openpyxl

REQUIRED_COLS = ["full name", "email", "company"]
OPTIONAL_COLS = ["position"]
EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


class ExcelParseError(Exception):
    pass


def parse_excel(path: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    # Build header map: {normalised_name: col_index}
    # Normalise: lowercase + collapse underscores/hyphens to spaces so that
    # "full_name", "full name", and "Full Name" all map to "full name".
    raw_headers = [str(cell.value or "").strip() for cell in ws[1]]
    header_map = {h.lower().replace("_", " ").replace("-", " "): i for i, h in enumerate(raw_headers)}

    # Validate required columns
    for col in REQUIRED_COLS:
        if col not in header_map:
            found = ", ".join(raw_headers) if raw_headers else "none"
            raise ExcelParseError(
                f"Column '{col}' not found. Found columns: {found} — "
                f"please rename to match expected headers."
            )

    def get(row, col_name):
        idx = header_map.get(col_name)
        if idx is None:
            return ""
        val = row[idx].value
        return str(val).strip() if val is not None else ""

    rows = []
    for row in ws.iter_rows(min_row=2):
        if all(cell.value is None for cell in row):
            continue  # skip blank rows
        full_name = get(row, "full name")
        email = get(row, "email")
        company = get(row, "company")
        position = get(row, "position")
        # Use explicit "First Name" column when present (template format);
        # fall back to splitting Full Name for simpler uploads.
        first_name = get(row, "first name") or (full_name.split()[0] if full_name else "")

        warning = None
        if not email:
            warning = "missing_email"
        elif not EMAIL_RE.match(email):
            warning = "invalid_email"

        rows.append({
            "full_name": full_name,
            "first_name": first_name,
            "email": email,
            "company": company,
            "position": position,
            "warning": warning,
        })

    # Flag duplicates — duplicate_name takes highest priority because the admin
    # must resolve ambiguous names before certificates can be generated.
    # Any pre-existing email warning is intentionally overridden.
    name_counts = Counter(r["full_name"] for r in rows)
    for row in rows:
        if name_counts[row["full_name"]] > 1:
            row["warning"] = "duplicate_name"

    return rows
