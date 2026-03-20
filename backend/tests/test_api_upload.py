import os
import pytest
import openpyxl
from fastapi.testclient import TestClient
from unittest.mock import patch, MagicMock

@pytest.fixture
def mock_supabase():
    mock = MagicMock()
    # workshops insert returns a workshop row
    mock.table.return_value.insert.return_value.execute.return_value.data = [
        {"id": "workshop-123", "name": "Test Workshop", "status": "draft", "created_at": "2026-01-01T00:00:00Z"}
    ]
    return mock

@pytest.fixture
def client(mock_supabase, tmp_path):
    with patch("db.client.get_supabase", return_value=mock_supabase):
        import importlib
        import sys
        # Clear any cached imports that may have already bound get_supabase
        for mod in list(sys.modules.keys()):
            if "api." in mod or mod in ("api.upload", "api.workshops", "api.participants", "api.generate", "api.email", "api.auth", "main"):
                sys.modules.pop(mod, None)
        from main import app
        return TestClient(app)

def make_test_xlsx(tmp_path):
    path = tmp_path / "test.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Full Name", "Email", "Company"])
    ws.append(["María García", "m@banco.com", "Banco"])
    wb.save(path)
    return path

def test_upload_valid_excel_returns_workshop_id(client, tmp_path):
    path = make_test_xlsx(tmp_path)
    with open(path, "rb") as f:
        response = client.post(
            "/api/upload",
            files={"file": ("test.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
    assert response.status_code == 200
    data = response.json()
    assert "workshop_id" in data
    assert data["participant_count"] == 1

def test_upload_returns_warning_count(client, tmp_path):
    path = tmp_path / "warn.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Full Name", "Email", "Company"])
    ws.append(["Test Person", "", "Co"])  # missing email
    wb.save(path)
    with open(path, "rb") as f:
        response = client.post(
            "/api/upload",
            files={"file": ("warn.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
    assert response.status_code == 200
    assert response.json()["warning_count"] == 1

def test_upload_rejects_missing_required_column(client, tmp_path):
    path = tmp_path / "bad.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Nombre", "Correo"])
    ws.append(["Test", "test@test.com"])
    wb.save(path)
    with open(path, "rb") as f:
        response = client.post(
            "/api/upload",
            files={"file": ("bad.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
    assert response.status_code == 422
    assert "found columns" in response.json()["detail"].lower()
